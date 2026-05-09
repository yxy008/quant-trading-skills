#!/usr/bin/env python3
"""
数据持久化模块 - SQLite存储回测结果、行情数据、交易记录
同时提供文件级K线缓存，加速数据获取
"""
import sqlite3
import json
import os
import time
import pickle
import hashlib
import threading
from datetime import datetime

DB_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data")
DB_PATH = os.path.join(DB_DIR, "quant.db")

os.makedirs(DB_DIR, exist_ok=True)

DB_LOCK = threading.Lock()


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


# ==================== 文件级K线缓存 ====================

CACHE_DIR = os.path.join(os.path.dirname(DB_DIR), ".cache")
CACHE_INDEX_FILE = os.path.join(CACHE_DIR, "cache_index.json")
CACHE_MAX_AGE_HOURS = 24


def _ensure_cache_dir():
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR)


def _get_cache_path(key):
    key_hash = hashlib.md5(key.encode('utf-8')).hexdigest()
    return os.path.join(CACHE_DIR, f"{key_hash}.pkl")


def _load_cache_index():
    _ensure_cache_dir()
    if os.path.exists(CACHE_INDEX_FILE):
        try:
            with open(CACHE_INDEX_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_cache_index(index):
    _ensure_cache_dir()
    try:
        with open(CACHE_INDEX_FILE, 'w', encoding='utf-8') as f:
            json.dump(index, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def get_cached_kline_file(symbol, days=250, adjust="qfq", max_age_hours=None):
    """
    从文件缓存获取K线DataFrame

    参数:
        symbol: 股票代码
        days: 数据天数
        adjust: 复权方式
        max_age_hours: 最大缓存时间（小时），默认24小时

    返回: DataFrame 或 None
    """
    if max_age_hours is None:
        max_age_hours = CACHE_MAX_AGE_HOURS

    cache_key = f"kline_{symbol}_{days}_{adjust}"
    cache_path = _get_cache_path(cache_key)
    index = _load_cache_index()

    if cache_key in index:
        entry = index[cache_key]
        cache_time = entry.get("时间", 0)
        age_hours = (time.time() - cache_time) / 3600

        if age_hours < max_age_hours and os.path.exists(cache_path):
            try:
                with open(cache_path, 'rb') as f:
                    return pickle.load(f)
            except Exception:
                pass

    return None


def save_cached_kline_file(symbol, days, adjust, df):
    """保存K线DataFrame到文件缓存"""
    _ensure_cache_dir()
    cache_key = f"kline_{symbol}_{days}_{adjust}"
    cache_path = _get_cache_path(cache_key)

    try:
        with open(cache_path, 'wb') as f:
            pickle.dump(df, f)

        index = _load_cache_index()
        index[cache_key] = {
            "类型": "K线",
            "股票代码": symbol,
            "天数": days,
            "复权": adjust,
            "时间": time.time(),
            "大小": os.path.getsize(cache_path),
        }
        _save_cache_index(index)
        return True
    except Exception:
        return False


def get_or_fetch_kline(symbol, days=250, adjust="qfq", force_refresh=False):
    """
    获取K线数据（优先文件缓存，未命中则从网络获取并缓存）

    参数:
        symbol: 股票代码
        days: 数据天数
        adjust: 复权方式
        force_refresh: 是否强制刷新

    返回: DataFrame 或 None
    """
    if not force_refresh:
        cached = get_cached_kline_file(symbol, days, adjust)
        if cached is not None:
            return cached

    try:
        from data_utils import get_stock_kline
        df = get_stock_kline(symbol, days, adjust)
    except ImportError:
        df = None

    if df is not None and len(df) > 0:
        save_cached_kline_file(symbol, days, adjust, df)

    return df


def preload_common_stocks(symbols=None, days=500):
    """
    预热缓存：预加载常用股票数据

    参数:
        symbols: 股票代码列表，默认加载常见蓝筹股
        days: 数据天数

    返回: 预热结果
    """
    if symbols is None:
        symbols = [
            "600519", "000858", "000568",
            "600036", "601318", "000001",
            "600276", "000651", "002415",
            "601012", "300750", "002594",
            "600900", "601088", "600585",
        ]

    results = []
    success = 0
    failed = 0

    for symbol in symbols:
        df = get_or_fetch_kline(symbol, days)
        if df is not None and len(df) > 0:
            success += 1
            results.append({
                "股票代码": symbol,
                "状态": "成功",
                "数据条数": len(df),
            })
        else:
            failed += 1
            results.append({
                "股票代码": symbol,
                "状态": "失败",
            })

    return {
        "预热时间": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "总计": len(symbols),
        "成功": success,
        "失败": failed,
        "详情": results,
    }


def get_file_cache_info():
    """获取文件缓存详细信息"""
    index = _load_cache_index()

    total_size = 0
    by_type = {}
    expired_count = 0

    now = time.time()
    for key, entry in index.items():
        size = entry.get("大小", 0)
        total_size += size

        cache_type = entry.get("类型", "未知")
        if cache_type not in by_type:
            by_type[cache_type] = {"数量": 0, "大小": 0}
        by_type[cache_type]["数量"] += 1
        by_type[cache_type]["大小"] += size

        age_hours = (now - entry.get("时间", 0)) / 3600
        if age_hours > CACHE_MAX_AGE_HOURS:
            expired_count += 1

    return {
        "缓存目录": CACHE_DIR,
        "总条目数": len(index),
        "总大小": f"{total_size / 1024 / 1024:.2f} MB",
        "过期条目数": expired_count,
        "最大保存时间": f"{CACHE_MAX_AGE_HOURS}小时",
        "按类型统计": by_type,
        "分析时间": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }


def clear_file_cache():
    """清理所有文件缓存"""
    _ensure_cache_dir()

    deleted_files = 0
    deleted_size = 0

    for filename in os.listdir(CACHE_DIR):
        filepath = os.path.join(CACHE_DIR, filename)
        if os.path.isfile(filepath):
            try:
                deleted_size += os.path.getsize(filepath)
                os.remove(filepath)
                deleted_files += 1
            except Exception:
                pass

    if os.path.exists(CACHE_INDEX_FILE):
        try:
            os.remove(CACHE_INDEX_FILE)
        except Exception:
            pass

    return {
        "清理文件数": deleted_files,
        "释放空间": f"{deleted_size / 1024 / 1024:.2f} MB",
        "清理时间": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }


def clear_expired_file_cache():
    """清理过期文件缓存"""
    index = _load_cache_index()
    now = time.time()
    expired_keys = []
    deleted_size = 0

    for key, entry in index.items():
        age_hours = (now - entry.get("时间", 0)) / 3600
        if age_hours > CACHE_MAX_AGE_HOURS:
            cache_path = _get_cache_path(key)
            if os.path.exists(cache_path):
                try:
                    deleted_size += os.path.getsize(cache_path)
                    os.remove(cache_path)
                except Exception:
                    pass
            expired_keys.append(key)

    for key in expired_keys:
        del index[key]

    _save_cache_index(index)

    return {
        "清理条目数": len(expired_keys),
        "释放空间": f"{deleted_size / 1024 / 1024:.2f} MB",
        "清理时间": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }


def init_db():
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.executescript("""
            CREATE TABLE IF NOT EXISTS backtest_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                symbol TEXT NOT NULL,
                strategy TEXT NOT NULL,
                backtest_time TEXT NOT NULL,
                total_return REAL,
                annual_return REAL,
                sharpe_ratio REAL,
                max_drawdown REAL,
                win_rate REAL,
                trade_count INTEGER,
                calmar_ratio REAL,
                volatility REAL,
                metrics_json TEXT,
                created_at TEXT DEFAULT (datetime('now', 'localtime'))
            );

            CREATE TABLE IF NOT EXISTS daily_kline (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                symbol TEXT NOT NULL,
                trade_date TEXT NOT NULL,
                open REAL,
                high REAL,
                low REAL,
                close REAL,
                volume REAL,
                amount REAL,
                created_at TEXT DEFAULT (datetime('now', 'localtime')),
                UNIQUE(symbol, trade_date)
            );

            CREATE TABLE IF NOT EXISTS trade_signals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                symbol TEXT NOT NULL,
                strategy TEXT NOT NULL,
                signal_date TEXT NOT NULL,
                signal_type TEXT NOT NULL,
                price REAL,
                reason TEXT,
                created_at TEXT DEFAULT (datetime('now', 'localtime'))
            );

            CREATE TABLE IF NOT EXISTS strategy_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                strategy_name TEXT NOT NULL UNIQUE,
                params_json TEXT,
                description TEXT,
                updated_at TEXT DEFAULT (datetime('now', 'localtime'))
            );

            CREATE INDEX IF NOT EXISTS idx_bt_symbol ON backtest_records(symbol);
            CREATE INDEX IF NOT EXISTS idx_bt_strategy ON backtest_records(strategy);
            CREATE INDEX IF NOT EXISTS idx_bt_time ON backtest_records(backtest_time);
            CREATE INDEX IF NOT EXISTS idx_kline_symbol ON daily_kline(symbol);
            CREATE INDEX IF NOT EXISTS idx_kline_date ON daily_kline(trade_date);
            CREATE INDEX IF NOT EXISTS idx_signal_symbol ON trade_signals(symbol);
            CREATE INDEX IF NOT EXISTS idx_signal_date ON trade_signals(signal_date);
        """)
        conn.commit()
        conn.close()


def save_backtest_record(symbol, strategy, metrics):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO backtest_records (symbol, strategy, backtest_time,
                total_return, annual_return, sharpe_ratio, max_drawdown,
                win_rate, trade_count, calmar_ratio, volatility, metrics_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            symbol,
            strategy,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            metrics.get('累计收益率'),
            metrics.get('年化收益率'),
            metrics.get('夏普比率'),
            metrics.get('最大回撤'),
            metrics.get('胜率'),
            metrics.get('交易次数'),
            metrics.get('Calmar比率'),
            metrics.get('年化波动率'),
            json.dumps(metrics, ensure_ascii=False, default=str)
        ))
        conn.commit()
        record_id = cursor.lastrowid
        conn.close()
        return record_id


def get_backtest_records(limit=100, symbol=None, strategy=None):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        query = "SELECT * FROM backtest_records WHERE 1=1"
        params = []
        if symbol:
            query += " AND symbol = ?"
            params.append(symbol)
        if strategy:
            query += " AND strategy = ?"
            params.append(strategy)
        query += " ORDER BY id DESC LIMIT ?"
        params.append(limit)
        cursor.execute(query, params)
        rows = cursor.fetchall()
        records = []
        for row in rows:
            records.append({
                "id": row["id"],
                "股票代码": row["symbol"],
                "策略": row["strategy"],
                "回测时间": row["backtest_time"],
                "累计收益率": row["total_return"],
                "年化收益率": row["annual_return"],
                "夏普比率": row["sharpe_ratio"],
                "最大回撤": row["max_drawdown"],
                "胜率": row["win_rate"],
                "交易次数": row["trade_count"],
                "Calmar比率": row["calmar_ratio"],
                "年化波动率": row["volatility"]
            })
        conn.close()
        return records


def delete_backtest_record(record_id):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM backtest_records WHERE id = ?", (record_id,))
        conn.commit()
        affected = cursor.rowcount
        conn.close()
        return affected > 0


def clear_backtest_records():
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM backtest_records")
        conn.commit()
        conn.close()


def save_kline_data(symbol, df):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        for _, row in df.iterrows():
            trade_date = str(row.get('date', row.get('日期', '')))
            if not trade_date:
                continue
            cursor.execute("""
                INSERT OR REPLACE INTO daily_kline (symbol, trade_date, open, high, low, close, volume, amount)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                symbol,
                trade_date[:10] if len(trade_date) > 10 else trade_date,
                float(row.get('open', row.get('开盘', 0))),
                float(row.get('high', row.get('最高', 0))),
                float(row.get('low', row.get('最低', 0))),
                float(row.get('close', row.get('收盘', 0))),
                float(row.get('volume', row.get('成交量', 0))),
                float(row.get('amount', row.get('成交额', 0))) if ('amount' in row or '成交额' in row) else 0
            ))
        conn.commit()
        conn.close()


def get_cached_kline(symbol, days=250):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM daily_kline WHERE symbol = ?
            ORDER BY trade_date DESC LIMIT ?
        """, (symbol, days))
        rows = cursor.fetchall()
        conn.close()
        return [dict(row) for row in rows]


def save_trade_signal(symbol, strategy, signal_date, signal_type, price, reason=""):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO trade_signals (symbol, strategy, signal_date, signal_type, price, reason)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (symbol, strategy, signal_date, signal_type, price, reason))
        conn.commit()
        conn.close()


def get_recent_signals(symbol=None, days=30):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        query = "SELECT * FROM trade_signals WHERE 1=1"
        params = []
        if symbol:
            query += " AND symbol = ?"
            params.append(symbol)
        query += " ORDER BY signal_date DESC LIMIT 100"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        return [dict(row) for row in rows]


def save_strategy_config(strategy_name, params, description=""):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT OR REPLACE INTO strategy_config (strategy_name, params_json, description, updated_at)
            VALUES (?, ?, ?, ?)
        """, (
            strategy_name,
            json.dumps(params, ensure_ascii=False),
            description,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ))
        conn.commit()
        conn.close()


def get_strategy_config(strategy_name=None):
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        if strategy_name:
            cursor.execute("SELECT * FROM strategy_config WHERE strategy_name = ?", (strategy_name,))
        else:
            cursor.execute("SELECT * FROM strategy_config ORDER BY updated_at DESC")
        rows = cursor.fetchall()
        conn.close()
        return [dict(row) for row in rows]


def get_backtest_trend(symbol, strategy, limit=20):
    """
    获取同一策略在同一股票上的回测趋势
    按时间排序，展示收益率、夏普比率等指标的变化趋势
    """
    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM backtest_records
            WHERE symbol = ? AND strategy = ?
            ORDER BY created_at ASC
            LIMIT ?
        """, (symbol, strategy, limit))
        rows = cursor.fetchall()
        conn.close()

        if not rows:
            return {"error": f"未找到 {symbol} 使用 {strategy} 的回测记录"}

        trend_data = []
        for row in rows:
            trend_data.append({
                "回测时间": row["backtest_time"],
                "累计收益率": row["total_return"],
                "年化收益率": row["annual_return"],
                "夏普比率": row["sharpe_ratio"],
                "最大回撤": row["max_drawdown"],
                "胜率": row["win_rate"],
                "交易次数": row["trade_count"],
                "Calmar比率": row["calmar_ratio"],
                "年化波动率": row["volatility"],
            })

        first = trend_data[0]
        last = trend_data[-1]

        changes = {}
        for key in ["累计收益率", "年化收益率", "夏普比率", "最大回撤", "胜率"]:
            if first.get(key) is not None and last.get(key) is not None:
                changes[key] = {
                    "首次": first[key],
                    "最新": last[key],
                    "变化": round(last[key] - first[key], 4) if isinstance(last[key], (int, float)) else None,
                }

        return {
            "股票代码": symbol,
            "策略": strategy,
            "记录数": len(trend_data),
            "趋势数据": trend_data,
            "指标变化": changes,
            "趋势判断": _analyze_trend(changes),
        }


def _analyze_trend(changes):
    """分析回测指标变化趋势"""
    analysis = []

    ret_change = changes.get("累计收益率", {}).get("变化")
    if ret_change is not None:
        if ret_change > 5:
            analysis.append("收益率持续提升，策略表现改善")
        elif ret_change < -5:
            analysis.append("收益率持续下降，策略表现恶化，需关注")
        else:
            analysis.append("收益率基本稳定")

    sharpe_change = changes.get("夏普比率", {}).get("变化")
    if sharpe_change is not None:
        if sharpe_change > 0.3:
            analysis.append("风险调整收益改善")
        elif sharpe_change < -0.3:
            analysis.append("风险调整收益下降")

    dd_change = changes.get("最大回撤", {}).get("变化")
    if dd_change is not None:
        if dd_change < -2:
            analysis.append("最大回撤收窄，风控改善")
        elif dd_change > 2:
            analysis.append("最大回撤扩大，风险增加")

    wr_change = changes.get("胜率", {}).get("变化")
    if wr_change is not None:
        if wr_change > 5:
            analysis.append("胜率提升，信号质量改善")
        elif wr_change < -5:
            analysis.append("胜率下降，信号质量恶化")

    return analysis if analysis else ["指标变化不显著，策略表现稳定"]


def compare_backtest_strategies(symbol, strategy_ids=None, limit_per_strategy=5):
    """
    对比多个策略在同一股票上的回测趋势
    """
    if strategy_ids is None:
        strategy_ids = []

    with DB_LOCK:
        conn = get_connection()
        cursor = conn.cursor()

        results = {}
        for sid in strategy_ids:
            cursor.execute("""
                SELECT * FROM backtest_records
                WHERE symbol = ? AND strategy = ?
                ORDER BY created_at DESC
                LIMIT ?
            """, (symbol, sid, limit_per_strategy))
            rows = cursor.fetchall()
            if rows:
                latest = rows[0]
                results[sid] = {
                    "策略": sid,
                    "最新回测时间": latest["backtest_time"],
                    "累计收益率": latest["total_return"],
                    "年化收益率": latest["annual_return"],
                    "夏普比率": latest["sharpe_ratio"],
                    "最大回撤": latest["max_drawdown"],
                    "胜率": latest["win_rate"],
                    "交易次数": latest["trade_count"],
                    "记录数": len(rows),
                }

        conn.close()

        sorted_results = sorted(
            results.values(),
            key=lambda x: x.get("夏普比率", 0) or 0,
            reverse=True
        )

        return {
            "股票代码": symbol,
            "策略数量": len(sorted_results),
            "策略对比": sorted_results,
            "最佳策略": sorted_results[0] if sorted_results else None,
    }


def export_to_excel(backtest_result, filepath=None):
    """
    将回测结果导出为Excel文件
    返回: (filepath, bytes) 或 (None, error_msg)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, "请安装openpyxl库: pip install openpyxl"

    try:
        import pandas as pd
    except ImportError:
        return None, "请安装pandas库: pip install pandas"

    try:
        wb = openpyxl.Workbook()

        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        def style_header(ws, row, cols):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def style_cell(ws, row, col, value=None):
            cell = ws.cell(row=row, column=col)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if value is not None:
                cell.value = value
            return cell

        ws1 = wb.active
        ws1.title = "绩效指标"

        metrics = [
            ("策略名称", backtest_result.get("策略名称", "")),
            ("股票代码", backtest_result.get("股票代码", "")),
            ("初始资金", backtest_result.get("初始资金", 0)),
            ("最终价值", backtest_result.get("最终价值", 0)),
            ("总收益率(%)", backtest_result.get("总收益率", 0)),
            ("年化收益率(%)", backtest_result.get("年化收益率", 0)),
            ("年化波动率(%)", backtest_result.get("年化波动率", 0)),
            ("夏普比率", backtest_result.get("夏普比率", 0)),
            ("最大回撤(%)", backtest_result.get("最大回撤", 0)),
            ("胜率(%)", backtest_result.get("胜率", 0)),
            ("盈亏比", backtest_result.get("盈亏比", 0)),
            ("交易次数", backtest_result.get("交易次数", 0)),
            ("盈利次数", backtest_result.get("盈利次数", 0)),
            ("亏损次数", backtest_result.get("亏损次数", 0)),
            ("平均盈利(%)", backtest_result.get("平均盈利", 0)),
            ("平均亏损(%)", backtest_result.get("平均亏损", 0)),
            ("最大连续盈利", backtest_result.get("最大连续盈利", 0)),
            ("最大连续亏损", backtest_result.get("最大连续亏损", 0)),
            ("回测天数", backtest_result.get("回测天数", 0)),
            ("数据起始", str(backtest_result.get("数据起始", ""))),
            ("数据截止", str(backtest_result.get("数据截止", ""))),
            ("佣金费率", backtest_result.get("佣金费率", 0)),
            ("滑点比例", backtest_result.get("滑点比例", 0)),
        ]

        ws1.column_dimensions['A'].width = 18
        ws1.column_dimensions['B'].width = 20

        for i, (label, value) in enumerate(metrics, 1):
            style_cell(ws1, i, 1, label)
            style_cell(ws1, i, 2, value)

        ws2 = wb.create_sheet("交易记录")
        trades = backtest_result.get("交易记录", [])

        headers = ["序号", "日期", "类型", "价格", "数量", "金额", "手续费", "盈亏", "盈亏比例(%)", "持仓市值"]
        for col, h in enumerate(headers, 1):
            style_cell(ws2, 1, col, h)
        style_header(ws2, 1, len(headers))

        for i, trade in enumerate(trades):
            row = i + 2
            style_cell(ws2, row, 1, i + 1)
            style_cell(ws2, row, 2, str(trade.get("日期", "")))
            style_cell(ws2, row, 3, trade.get("类型", ""))
            style_cell(ws2, row, 4, trade.get("价格", 0))
            style_cell(ws2, row, 5, trade.get("数量", 0))
            style_cell(ws2, row, 6, trade.get("金额", 0))
            style_cell(ws2, row, 7, trade.get("手续费", 0))
            style_cell(ws2, row, 8, trade.get("盈亏", 0))
            style_cell(ws2, row, 9, trade.get("盈亏比例", 0))
            style_cell(ws2, row, 10, trade.get("持仓市值", 0))

        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 14

        ws3 = wb.create_sheet("权益曲线")
        equity = backtest_result.get("权益曲线", [])

        eq_headers = ["日期", "权益", "收益率(%)", "回撤(%)", "持仓方向"]
        for col, h in enumerate(eq_headers, 1):
            style_cell(ws3, 1, col, h)
        style_header(ws3, 1, len(eq_headers))

        for i, eq in enumerate(equity):
            row = i + 2
            style_cell(ws3, row, 1, str(eq.get("日期", "")))
            style_cell(ws3, row, 2, eq.get("权益", 0))
            style_cell(ws3, row, 3, eq.get("收益率", 0))
            style_cell(ws3, row, 4, eq.get("回撤", 0))
            style_cell(ws3, row, 5, eq.get("持仓方向", "空仓"))

        for col in range(1, len(eq_headers) + 1):
            ws3.column_dimensions[get_column_letter(col)].width = 16

        if filepath:
            wb.save(filepath)
            return filepath, None
        else:
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return None, output.getvalue()

    except Exception as e:
        return None, f"导出Excel失败: {str(e)}"


def export_to_html(backtest_result):
    """
    将回测结果导出为HTML报告
    """
    strategy_name = backtest_result.get("策略名称", "未知策略")
    symbol = backtest_result.get("股票代码", "")
    total_return = backtest_result.get("总收益率", 0)
    annual_return = backtest_result.get("年化收益率", 0)
    sharpe = backtest_result.get("夏普比率", 0)
    max_dd = backtest_result.get("最大回撤", 0)
    win_rate = backtest_result.get("胜率", 0)
    trade_count = backtest_result.get("交易次数", 0)
    start_date = backtest_result.get("数据起始", "")
    end_date = backtest_result.get("数据截止", "")

    return_color = "#CC0000" if total_return >= 0 else "#008000"
    dd_color = "#CC0000"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>回测报告 - {strategy_name} - {symbol}</title>
<style>
    * {{ margin: 0; padding: 0; box-sizing: border-box; }}
    body {{ font-family: 'Microsoft YaHei', '微软雅黑', sans-serif; background: #f5f6fa; color: #333; padding: 20px; }}
    .container {{ max-width: 1000px; margin: 0 auto; }}
    .header {{ background: linear-gradient(135deg, #2B579A, #1a3a6e); color: #fff; padding: 30px; border-radius: 12px; margin-bottom: 20px; }}
    .header h1 {{ font-size: 24px; margin-bottom: 8px; }}
    .header p {{ font-size: 14px; opacity: 0.85; }}
    .card {{ background: #fff; border-radius: 10px; padding: 24px; margin-bottom: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
    .card h2 {{ font-size: 18px; margin-bottom: 16px; color: #2B579A; border-bottom: 2px solid #e8ecf1; padding-bottom: 8px; }}
    .metrics {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 12px; }}
    .metric {{ padding: 12px; background: #f8f9fc; border-radius: 8px; }}
    .metric-label {{ font-size: 12px; color: #888; margin-bottom: 4px; }}
    .metric-value {{ font-size: 20px; font-weight: 700; }}
    .positive {{ color: #CC0000; }}
    .negative {{ color: #008000; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th {{ background: #2B579A; color: #fff; padding: 10px 12px; text-align: center; font-weight: 600; }}
    td {{ padding: 8px 12px; text-align: center; border-bottom: 1px solid #eee; }}
    tr:hover td {{ background: #f5f7fb; }}
    .footer {{ text-align: center; color: #999; font-size: 12px; margin-top: 20px; padding: 16px; }}
    .buy {{ color: #CC0000; font-weight: 600; }}
    .sell {{ color: #008000; font-weight: 600; }}
</style>
</head>
<body>
<div class="container">
    <div class="header">
        <h1>策略回测报告</h1>
        <p>{strategy_name} | {symbol} | 回测区间: {start_date} ~ {end_date}</p>
        <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>

    <div class="card">
        <h2>核心绩效指标</h2>
        <div class="metrics">
            <div class="metric">
                <div class="metric-label">总收益率</div>
                <div class="metric-value" style="color:{return_color}">{total_return:.2f}%</div>
            </div>
            <div class="metric">
                <div class="metric-label">年化收益率</div>
                <div class="metric-value" style="color:{return_color}">{annual_return:.2f}%</div>
            </div>
            <div class="metric">
                <div class="metric-label">夏普比率</div>
                <div class="metric-value">{sharpe:.2f}</div>
            </div>
            <div class="metric">
                <div class="metric-label">最大回撤</div>
                <div class="metric-value" style="color:{dd_color}">{max_dd:.2f}%</div>
            </div>
            <div class="metric">
                <div class="metric-label">胜率</div>
                <div class="metric-value">{win_rate:.1f}%</div>
            </div>
            <div class="metric">
                <div class="metric-label">交易次数</div>
                <div class="metric-value">{trade_count}</div>
            </div>
            <div class="metric">
                <div class="metric-label">初始资金</div>
                <div class="metric-value">{backtest_result.get('初始资金', 0):,.0f}</div>
            </div>
            <div class="metric">
                <div class="metric-label">最终价值</div>
                <div class="metric-value">{backtest_result.get('最终价值', 0):,.0f}</div>
            </div>
        </div>
    </div>

    <div class="card">
        <h2>交易记录</h2>
        <table>
            <thead>
                <tr>
                    <th>序号</th><th>日期</th><th>类型</th><th>价格</th>
                    <th>数量</th><th>金额</th><th>手续费</th><th>盈亏</th>
                </tr>
            </thead>
            <tbody>
"""

    trades = backtest_result.get("交易记录", [])
    for i, t in enumerate(trades):
        trade_type = t.get("类型", "")
        type_cls = "buy" if "买入" in trade_type else "sell"
        pnl = t.get("盈亏", 0)
        pnl_cls = "positive" if pnl > 0 else ("negative" if pnl < 0 else "")
        html += f"""<tr>
            <td>{i + 1}</td>
            <td>{t.get('日期', '')}</td>
            <td class="{type_cls}">{trade_type}</td>
            <td>{t.get('价格', 0):.2f}</td>
            <td>{t.get('数量', 0)}</td>
            <td>{t.get('金额', 0):,.0f}</td>
            <td>{t.get('手续费', 0):.2f}</td>
            <td class="{pnl_cls}">{pnl:+,.2f}</td>
        </tr>"""

    html += """
            </tbody>
        </table>
    </div>

    <div class="footer">
        <p>本报告由量化交易系统自动生成 | 仅供参考，不构成投资建议</p>
    </div>
</div>
</body>
</html>"""

    return html


init_db()
